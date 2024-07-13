import tkinter as tk
import os
import _Launch as launch
from tkinter import messagebox as mb
from openpyxl import load_workbook


class MainGui:
    def __init__(self):
        self.root = tk.Tk()
        #self.root.iconbitmap(default="appicon.ico")

        def center_window(window, width, height):
            # Get screen width and height
            screen_width = window.winfo_screenwidth()
            screen_height = window.winfo_screenheight()

            # Calculate position x and y
            x = (screen_width // 2) - (width // 2)
            y = (screen_height // 2) - (height// 2)

            # Set the window position
            window.geometry(f'{width}x{height}+{x}+{y}')



        # Set window size
        window_width = 1500
        window_height = 800
        center_window(self.root, window_width, window_height)



        self.root.title("Bulk URL Opener")
        obj = EditXLSX(self.root)
        self.root.mainloop()


class EditXLSX:
    def __init__(self, root):
        self.root = root
        self.frameSelections = tk.Frame(self.root)
        self.frameSelections.pack()
        tk.Label(self.frameSelections, text="Select your Bookmarks:",font=("times new Roman",29)).pack()
        self.list = os.listdir(r"Bookmarks/")

        tk.Label(self.frameSelections,text="Bookmarks:",font=("Times New Roman",20)).pack(side=tk.LEFT)
        self.selected_option_1 = tk.StringVar(self.root)
        self.selected_option_1.set(self.list[0])
        option_menu_1 = tk.OptionMenu(self.frameSelections,self.selected_option_1, *self.list,command=self.on_category_selected)
        option_menu_1.pack(pady=20,side = tk.LEFT)
        option_menu_1.config(font=("Times New Roman",18))
        self.link = "Bookmarks/{}".format(self.selected_option_1.get())
        # Set the default choice

        self.wb = load_workbook("Bookmarks/{}".format(self.selected_option_1.get()))
        self.choices = self.wb.sheetnames

        self.selected_option = tk.StringVar(self.root)
        self.selected_option.set(self.choices[0])
        self.option_menu = tk.OptionMenu(self.frameSelections, self.selected_option, *self.choices)
        self.option_menu.pack(pady=20,side = tk.LEFT)
        self.option_menu.config(font=("Times New Roman", 18))





        self.frameButtons = tk.Frame(self.root)
        self.frameButtons.pack()

        self.display_instance = display()
        #self.on_refresh()
        tk.Button(self.frameButtons, text="View Contents", command=self.on_refresh,font=("Times New Roman",20)).pack(side=tk.LEFT)
        tk.Button(self.frameButtons, text="Open in Browser", command=lambda: launch.Launch(self.selected_option.get(),sheet=self.link),font=("Times New Roman",20)).pack(side=tk.BOTTOM)
        self.root.mainloop()


    def on_category_selected(self,event):
        self.option = self.selected_option_1.get()
        self.link = "Bookmarks/{}".format(self.option)
        if os.path.exists(self.link):
            workbook = load_workbook(self.link)
            self.choices = workbook.sheetnames
            self.selected_option.set(self.choices[0])
            self.option_menu['menu'].delete(0, 'end')
            for i in self.choices:
                self.option_menu["menu"].add_command(label = i, command = tk._setit(self.selected_option, i))




    def on_refresh(self):
        if self.link.strip() != "":
            self.display_instance.clear_text()
            self.display_instance.display_excel_contents(file_path=self.link, sheet_name=self.selected_option.get(),root=self.root)



class display:
    def display_excel_contents(self, file_path, sheet_name, root):
        self.root = root
        self.root.title("Contents of : {}_{}".format(file_path,sheet_name))
        self.text_box = tk.Text(root, height=20, width=60,background="Grey",font=("Times new Roman",18),foreground="white")
        self.text_box.pack(padx=10, pady=10)
        try:
            wb = load_workbook(file_path)
            sheet = wb[sheet_name]

            for row in sheet.iter_rows():
                row_data = ""
                for cell in row:
                    row_data += str(cell.value).ljust(15)  # Adjust width as needed
                self.text_box.insert(tk.END, row_data + "\n")

        except FileNotFoundError:
            mb.showerror("Error", f"File '{file_path}' not found.")
        except KeyError:
            mb.showerror("Error", f"Sheet '{sheet_name}' not found in '{file_path}'.")

    def clear_text(self):
        if hasattr(self, 'text_box'):
            self.text_box.destroy()


