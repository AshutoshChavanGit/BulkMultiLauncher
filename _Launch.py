import webbrowser
from openpyxl import load_workbook

class Launch:
    def __init__(self,option,sheet):
        self.sheet = sheet
        self.option = option
        wb = load_workbook(self.sheet)
        self.choices = wb.sheetnames
        self.launcher()

    def launcher(self):
        wb1 = load_workbook(self.sheet)
        sheet = wb1[self.option]

        for cell in sheet['A']:
            if cell.value:
                webbrowser.open_new_tab(cell.value)
        wb1.close()

